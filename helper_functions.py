import json
import os
from openpyxl import load_workbook
from urllib.parse import urlparse, parse_qs
from collections import OrderedDict


def find_har_file():
    for file in os.listdir('.'):
        if file.endswith('.har'):
            print(f"Found HAR file: {file}")
            return file
    print("No HAR file found in the current directory.")
    return None


def load_har_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_tests_from_xlsx(xlsx_path):
    test_cases = []
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_case = {
            "name": row[0],
            "description": row[1],
            "target_urls": row[2].split(','),
            "parameter_checks": [{
                "name": row[3],
                "condition": row[4],
                "value": row[5],
                "optional": str(row[6]).lower() == 'true',
                "on_pass": row[7],
                "on_fail": row[8]
            }]
        }
        test_cases.append(test_case)
    return test_cases


def extract_parameters(url, post_data_params):
    parsed_url = urlparse(url)
    url_params = parse_qs(parsed_url.query)
    url_params = {key: value[0] for key, value in url_params.items()}
    combined_params = {**url_params, **post_data_params}
    return combined_params


def apply_test_cases(call, test_cases):
    results = []
    parsed_url = urlparse(call["url"])
    call_domain = parsed_url.netloc

    for test in test_cases:
        target_urls = test.get("target_urls", [])
        if target_urls and not any(target in call_domain for target in target_urls):
            continue

        conditions_met = True
        for condition in test.get("conditions", []):
            cond_name = condition["name"]
            cond_values = condition.get("value", [])
            is_optional = condition.get("optional", False)
            param_value = call["parameters"].get(cond_name)
            if param_value:
                if cond_values and param_value not in cond_values:
                    conditions_met = False
                    break
            elif not is_optional:
                conditions_met = False
                break

        if not conditions_met:
            continue

        for action in test.get("actions", []):
            if action["type"] == "flag":
                results.append(f"Flag: {action['message']}")

        for param_check in test.get("parameter_checks", []):
            param_names = param_check.get("names", [param_check.get("name")])
            condition = param_check.get("condition", "")
            found = False
            found_value = None
            for name in param_names:
                if name in call["parameters"]:
                    found = True
                    found_value = call["parameters"][name]
                    break

            if condition == "exists":
                if found:
                    results.append((param_check["on_pass"].format(value=found_value, url=call["url"]), param_check["name"]))
                    dependent_tests = test.get("dependent_tests", "")
                    if dependent_tests:
                        dependent_test_names = dependent_tests.split(',')
                        for dep_test_name in dependent_test_names:
                            dep_test = next((t for t in test_cases if t["name"] == dep_test_name.strip()), None)
                            if dep_test:
                                dep_results = apply_test_cases(call, [dep_test])
                                results.extend(dep_results)
                else:
                    results.append((param_check["on_fail"].format(url=call["url"]), param_check["name"]))

    return results


def parse_har_for_calls(har_data, test_cases):
    calls = []
    for entry in har_data['log']['entries']:
        url = entry['request']['url']
        post_data = entry['request'].get('postData', {})
        post_data_params = {param["name"]: param["value"] for param in post_data.get('params', [])}
        parameters = extract_parameters(url, post_data_params)

        call = {
            "url": url,
            "parameters": parameters,
            "payload": post_data.get("text", "No payload")
        }

        results = apply_test_cases(call, test_cases)
        call["results"] = results
        calls.append(call)
    return calls


def combine_nested_keys(params):
    combined_params = OrderedDict()
    current_key = ""

    for key, value in params.items():
        if key.endswith('.') or key.startswith('.'):
            current_key += key.strip('.')
        else:
            full_key = f"{current_key}.{key}" if current_key else key
            combined_params[full_key] = value
            current_key = ""

    return combined_params
