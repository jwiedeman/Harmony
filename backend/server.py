import os
import json
import uuid
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
import openpyxl
from collections import OrderedDict
from urllib.parse import urlparse, parse_qs
import tempfile
import shutil
import glob
from openpyxl import load_workbook

from .parsers.chlsj_parser import parse_chlsj
import io

app = FastAPI(title="Harmony QA API", description="QA Testing API for HAR file analysis")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def parse_har(file_obj: io.TextIOBase):
    """Convert a HAR file object into a list of network events.

    The schema matches the output of :func:`parse_chlsj` so the frontend can
    treat both sources uniformly.  Only a subset of fields is extracted.
    """
    data = json.load(file_obj)
    entries = data.get("log", {}).get("entries", [])
    events = []
    for entry in entries:
        request = entry.get("request", {})
        response = entry.get("response", {})
        event = {
            "url": request.get("url"),
            "method": request.get("method"),
            "status": response.get("status"),
            "startedDateTime": entry.get("startedDateTime"),
            "requestHeaders": {h.get("name"): h.get("value") for h in request.get("headers", [])},
            "responseHeaders": {h.get("name"): h.get("value") for h in response.get("headers", [])},
            "postData": request.get("postData"),
            "queryParams": {p.get("name"): p.get("value") for p in request.get("queryString", [])},
            "bodyJSON": None,
            "raw": entry,
        }
        post = event.get("postData") or {}
        text = post.get("text")
        if isinstance(text, str):
            try:
                event["bodyJSON"] = json.loads(text)
            except json.JSONDecodeError:
                pass
        events.append(event)
    return events


@app.post("/api/ingest")
async def ingest(file: UploadFile = File(...)):
    """Ingest a network log in HAR or Charles .chlsj format.

    The response is a simple list of normalized network events that can be
    further processed by Harmony.  This endpoint is intentionally lightweight
    and does not perform any Adobe-specific validation."""
    ext = os.path.splitext(file.filename)[1].lower()
    contents = await file.read()
    try:
        text = contents.decode("utf-8")
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Unable to decode upload as UTF-8")

    with io.StringIO(text) as f:
        if ext == ".chlsj":
            events = parse_chlsj(f)
        elif ext == ".har":
            events = parse_har(f)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")
    return {"events": events}

# Pydantic models
class TestCase(BaseModel):
    id: str
    name: str
    description: str
    target_urls: List[str]
    parameter_name: str
    condition: str
    expected_value: Optional[str] = None
    optional: bool = False
    on_pass_message: str
    on_fail_message: str

class TestResult(BaseModel):
    url: str
    parameter: str
    result: str
    details: str
    test_case_name: str

class TestGroup(BaseModel):
    id: str
    name: str
    sequence: List[str]
    within_seconds: Optional[int] = None
    on_pass_message: Optional[str] = "Group {name} passed"
    on_fail_message: Optional[str] = "Group {name} failed"

class AnalysisReport(BaseModel):
    total_requests: int
    total_tests: int
    passed_tests: int
    failed_tests: int
    url_failures: Dict[str, int]
    dimension_failures: Dict[str, int]
    detailed_results: List[TestResult]
    raw_data: List[Dict[str, Any]]
    group_results: Dict[str, Dict[str, Any]]

class HarFile(BaseModel):
    filename: str
    path: str
    size: int
    modified: str

# In-memory storage for analysis results
analysis_results_store = {}

# Function to load test cases from Excel file
def load_tests_from_xlsx(xlsx_path='test_cases.xlsx'):
    """Load test cases from Excel file, similar to the original helper function."""
    test_cases = []
    try:
        # If path doesn't exist, try looking in parent directory
        if not os.path.exists(xlsx_path):
            parent_path = os.path.join("../", os.path.basename(xlsx_path))
            if os.path.exists(parent_path):
                xlsx_path = parent_path
        
        workbook = load_workbook(xlsx_path)
        sheet = workbook.active
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
                
            # Handle None values and convert to appropriate types
            target_urls = row[2].split(',') if row[2] else []
            target_urls = [url.strip() for url in target_urls if url.strip()]
            
            test_case = TestCase(
                id=str(uuid.uuid4()),
                name=str(row[0]) if row[0] else f"Test Case {row_num}",
                description=str(row[1]) if row[1] else "",
                target_urls=target_urls,
                parameter_name=str(row[3]) if row[3] else "",
                condition=str(row[4]).lower() if row[4] else "exists",
                expected_value=str(row[5]) if row[5] and row[5] != "None" else None,
                optional=str(row[6]).lower() == 'true' if row[6] else False,
                on_pass_message=str(row[7]) if row[7] else f"Parameter {row[3]} found: {{value}}",
                on_fail_message=str(row[8]) if row[8] else f"Parameter {row[3]} missing from {{url}}"
            )
            test_cases.append(test_case)
            
        print(f"Loaded {len(test_cases)} test cases from {xlsx_path}")
        return test_cases
        
    except FileNotFoundError:
        print(f"Test cases file {xlsx_path} not found. Using empty test case list.")
        return []
    except Exception as e:
        print(f"Error loading test cases from {xlsx_path}: {e}")
        return []

# Function to save test cases to Excel file
def save_tests_to_xlsx(test_cases: List[TestCase], xlsx_path='test_cases.xlsx'):
    """Save test cases to Excel file."""
    try:
        # If path doesn't exist, try parent directory
        if not os.path.exists(os.path.dirname(xlsx_path)) and xlsx_path != os.path.basename(xlsx_path):
            parent_path = os.path.join("../", os.path.basename(xlsx_path))
            xlsx_path = parent_path
        
        # Create new workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Test Cases"
        
        # Headers
        headers = [
            "Test Name", "Description", "Target URLs", "Parameter Name", 
            "Condition", "Expected Value", "Optional", "On Pass Message", "On Fail Message"
        ]
        
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, test_case in enumerate(test_cases, 2):
            sheet.cell(row=row, column=1, value=test_case.name)
            sheet.cell(row=row, column=2, value=test_case.description)
            sheet.cell(row=row, column=3, value=', '.join(test_case.target_urls))
            sheet.cell(row=row, column=4, value=test_case.parameter_name)
            sheet.cell(row=row, column=5, value=test_case.condition)
            sheet.cell(row=row, column=6, value=test_case.expected_value or "")
            sheet.cell(row=row, column=7, value=str(test_case.optional))
            sheet.cell(row=row, column=8, value=test_case.on_pass_message)
            sheet.cell(row=row, column=9, value=test_case.on_fail_message)
        
        workbook.save(xlsx_path)
        print(f"Saved {len(test_cases)} test cases to {xlsx_path}")
        return True
        
    except Exception as e:
        print(f"Error saving test cases to {xlsx_path}: {e}")
        return False

# Function to load test groups from JSON file
def load_test_groups_from_json(json_path='test_groups.json') -> List[TestGroup]:
    groups = []
    try:
        if not os.path.exists(json_path):
            parent_path = os.path.join("../", os.path.basename(json_path))
            if os.path.exists(parent_path):
                json_path = parent_path
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            for group in data:
                groups.append(TestGroup(
                    id=group.get('id', str(uuid.uuid4())),
                    name=group.get('name', ''),
                    sequence=group.get('sequence', []),
                    within_seconds=group.get('within_seconds'),
                    on_pass_message=group.get('on_pass_message', 'Group {name} passed'),
                    on_fail_message=group.get('on_fail_message', 'Group {name} failed')
                ))
    except FileNotFoundError:
        print(f"Test group file {json_path} not found. Continuing without groups.")
    except Exception as e:
        print(f"Error loading test groups from {json_path}: {e}")
    return groups

# Function to save test groups to JSON file
def save_test_groups_to_json(groups: List[TestGroup], json_path='test_groups.json') -> bool:
    try:
        # Ensure path
        if not os.path.exists(os.path.dirname(json_path)) and json_path != os.path.basename(json_path):
            parent_path = os.path.join("../", os.path.basename(json_path))
            json_path = parent_path

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump([g.dict() for g in groups], f, indent=2)
        print(f"Saved {len(groups)} test groups to {json_path}")
        return True
    except Exception as e:
        print(f"Error saving test groups to {json_path}: {e}")
        return False

# Function to get available HAR files
def get_available_har_files():
    """Get list of available HAR files in the project directory."""
    har_files = []
    
    # Look for HAR files in parent directory (project root) and common subdirectories  
    base_paths = [
        "../",  # Project root
        "../data/", 
        "../samples/",
        "../test_data/",
        "./",  # Current directory
        "data/",
        "samples/", 
        "test_data/"
    ]
    
    for base_path in base_paths:
        try:
            search_pattern = os.path.join(base_path, "*.har")
            for file_path in glob.glob(search_pattern):
                try:
                    stat = os.stat(file_path)
                    # Get relative path from project root
                    rel_path = os.path.relpath(file_path, "../")
                    har_files.append(HarFile(
                        filename=os.path.basename(file_path),
                        path=rel_path,
                        size=stat.st_size,
                        modified=str(stat.st_mtime)
                    ))
                except OSError:
                    continue
        except:
            continue
    
    # Remove duplicates based on filename
    seen_files = set()
    unique_har_files = []
    for har_file in har_files:
        if har_file.filename not in seen_files:
            unique_har_files.append(har_file)
            seen_files.add(har_file.filename)
    
    return unique_har_files

# Helper functions (ported from original helper_functions.py)
def extract_parameters(url: str, post_data_params: dict) -> dict:
    parsed_url = urlparse(url)
    url_params = parse_qs(parsed_url.query)
    url_params = {key: value[0] for key, value in url_params.items()}
    combined_params = {**url_params, **post_data_params}
    return combined_params

def apply_test_cases(call: dict, test_cases: List[TestCase]) -> List[tuple]:
    results = []
    parsed_url = urlparse(call["url"])
    call_domain = parsed_url.netloc
    
    for test in test_cases:
        # Check if this test applies to this URL
        if test.target_urls and not any(target in call_domain or target in call["url"] for target in test.target_urls):
            continue
        
        param_value = call["parameters"].get(test.parameter_name)
        
        if test.condition == "exists":
            if param_value is not None:
                result_msg = test.on_pass_message.format(value=param_value, url=call["url"])
                results.append((f"Pass: {result_msg}", test.parameter_name, test.name))
            else:
                if not test.optional:
                    result_msg = test.on_fail_message.format(url=call["url"])
                    results.append((f"Fail: {result_msg}", test.parameter_name, test.name))
        
        elif test.condition == "equals":
            if param_value is not None:
                if str(param_value) == str(test.expected_value):
                    result_msg = test.on_pass_message.format(value=param_value, url=call["url"])
                    results.append((f"Pass: {result_msg}", test.parameter_name, test.name))
                else:
                    result_msg = test.on_fail_message.format(url=call["url"])
                    results.append((f"Fail: Expected '{test.expected_value}' but got '{param_value}'", test.parameter_name, test.name))
            else:
                if not test.optional:
                    result_msg = test.on_fail_message.format(url=call["url"])
                    results.append((f"Fail: {result_msg}", test.parameter_name, test.name))
    
    return results

def parse_har_for_calls(har_data: dict, test_cases: List[TestCase]) -> List[dict]:
    calls = []
    for entry in har_data['log']['entries']:
        url = entry['request']['url']
        post_data = entry['request'].get('postData', {})
        
        # Parse POST data parameters
        post_data_params = {}
        if 'params' in post_data:
            post_data_params = {param["name"]: param["value"] for param in post_data.get('params', [])}
        elif 'text' in post_data:
            try:
                # Try to parse JSON from post data text
                json_data = json.loads(post_data['text'])
                if isinstance(json_data, dict):
                    post_data_params = json_data
            except:
                pass
        
        parameters = extract_parameters(url, post_data_params)
        
        call = {
            "url": url,
            "method": entry['request']['method'],
            "parameters": parameters,
            "payload": post_data.get("text", "No payload"),
            "status": entry['response']['status'],
            "startedDateTime": entry.get('startedDateTime')
        }
        
        results = apply_test_cases(call, test_cases)
        call["results"] = results
        calls.append(call)
    
    return calls

def analyze_failures(calls: List[dict]) -> tuple:
    url_failures = {}
    dimension_failures = {}
    
    for call in calls:
        for result, param_name, test_name in call.get("results", []):
            if "Fail" in result:
                url_failures[call['url']] = url_failures.get(call['url'], 0) + 1
                dimension_failures[param_name] = dimension_failures.get(param_name, 0) + 1
    
    return url_failures, dimension_failures

# Evaluate test groups based on call results
def evaluate_test_groups(calls: List[dict], groups: List[TestGroup]) -> Dict[str, Dict[str, Any]]:
    results: Dict[str, Dict[str, Any]] = {}
    # Map test names to list of indices when they passed
    test_pass_indices: Dict[str, List[int]] = {}
    for idx, call in enumerate(calls):
        for result, _, test_name in call.get("results", []):
            if "Pass" in result:
                test_pass_indices.setdefault(test_name, []).append(idx)

    for group in groups:
        current_index = -1
        group_passed = True
        first_time = None
        for test_name in group.sequence:
            indices = test_pass_indices.get(test_name, [])
            next_index = next((i for i in indices if i > current_index), None)
            if next_index is None:
                group_passed = False
                break
            if first_time is None:
                first_time = calls[next_index].get("startedDateTime")
            elif group.within_seconds is not None and first_time is not None:
                this_time = calls[next_index].get("startedDateTime")
                if this_time and first_time:
                    try:
                        from datetime import datetime
                        t1 = datetime.fromisoformat(first_time)
                        t2 = datetime.fromisoformat(this_time)
                        if (t2 - t1).total_seconds() > group.within_seconds:
                            group_passed = False
                            break
                    except Exception:
                        pass
            current_index = next_index
        message_template = group.on_pass_message if group_passed else group.on_fail_message
        message = message_template.format(name=group.name)
        results[group.name] = {"passed": group_passed, "message": message}
    return results

# API Endpoints
@app.get("/api/health")
async def health_check():
    return {"status": "healthy", "service": "Harmony QA API"}

@app.get("/api/test-cases")
async def get_test_cases():
    test_cases = load_tests_from_xlsx('../test_cases.xlsx')
    return {"test_cases": [tc.dict() for tc in test_cases]}

@app.post("/api/test-cases")
async def create_test_case(test_case: TestCase):
    # Load existing test cases
    test_cases = load_tests_from_xlsx('../test_cases.xlsx')
    
    # Add new test case
    if not test_case.id:
        test_case.id = str(uuid.uuid4())
    
    test_cases.append(test_case)
    
    # Save back to Excel
    if save_tests_to_xlsx(test_cases, '../test_cases.xlsx'):
        return {"message": "Test case created successfully", "id": test_case.id}
    else:
        raise HTTPException(status_code=500, detail="Failed to save test case")

@app.put("/api/test-cases/{test_case_id}")
async def update_test_case(test_case_id: str, test_case: TestCase):
    # Load existing test cases
    test_cases = load_tests_from_xlsx('../test_cases.xlsx')
    
    # Find and update the test case
    updated = False
    for i, tc in enumerate(test_cases):
        if tc.id == test_case_id:
            test_case.id = test_case_id
            test_cases[i] = test_case
            updated = True
            break
    
    if not updated:
        raise HTTPException(status_code=404, detail="Test case not found")
    
    # Save back to Excel
    if save_tests_to_xlsx(test_cases, '../test_cases.xlsx'):
        return {"message": "Test case updated successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to update test case")

@app.delete("/api/test-cases/{test_case_id}")
async def delete_test_case(test_case_id: str):
    # Load existing test cases
    test_cases = load_tests_from_xlsx('../test_cases.xlsx')
    
    # Find and remove the test case
    test_cases = [tc for tc in test_cases if tc.id != test_case_id]
    
    # Save back to Excel
    if save_tests_to_xlsx(test_cases, '../test_cases.xlsx'):
        return {"message": "Test case deleted successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to delete test case")

# Test Group Endpoints
@app.get("/api/test-groups")
async def get_test_groups():
    groups = load_test_groups_from_json('../test_groups.json')
    return {"test_groups": [g.dict() for g in groups]}


@app.post("/api/test-groups")
async def create_test_group(group: TestGroup):
    groups = load_test_groups_from_json('../test_groups.json')
    if not group.id:
        group.id = str(uuid.uuid4())
    groups.append(group)
    if save_test_groups_to_json(groups, '../test_groups.json'):
        return {"message": "Test group created successfully", "id": group.id}
    else:
        raise HTTPException(status_code=500, detail="Failed to save test group")


@app.put("/api/test-groups/{group_id}")
async def update_test_group(group_id: str, group: TestGroup):
    groups = load_test_groups_from_json('../test_groups.json')
    updated = False
    for i, g in enumerate(groups):
        if g.id == group_id:
            group.id = group_id
            groups[i] = group
            updated = True
            break
    if not updated:
        raise HTTPException(status_code=404, detail="Test group not found")
    if save_test_groups_to_json(groups, '../test_groups.json'):
        return {"message": "Test group updated successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to update test group")


@app.delete("/api/test-groups/{group_id}")
async def delete_test_group(group_id: str):
    groups = load_test_groups_from_json('../test_groups.json')
    groups = [g for g in groups if g.id != group_id]
    if save_test_groups_to_json(groups, '../test_groups.json'):
        return {"message": "Test group deleted successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to delete test group")

@app.get("/api/har-files")
async def get_har_files():
    har_files = get_available_har_files()
    return {"har_files": [hf.dict() for hf in har_files]}

@app.post("/api/analyze-har-file/{filename}")
async def analyze_har_file_by_name(filename: str):
    """Analyze a HAR file that exists in the project directory."""
    try:
        # Find the HAR file
        har_files = get_available_har_files()
        har_file = next((hf for hf in har_files if hf.filename == filename), None)
        
        if not har_file:
            raise HTTPException(status_code=404, detail=f"HAR file '{filename}' not found")
        
        # Construct full path - check if it's already a relative path from project root
        if har_file.path.startswith('../'):
            full_path = har_file.path
        else:
            full_path = os.path.join("../", har_file.path)
        
        # Read and parse HAR file
        with open(full_path, 'r', encoding='utf-8') as f:
            har_data = json.load(f)
        
        # Get test cases from Excel
        test_cases = load_tests_from_xlsx('../test_cases.xlsx')
        if not test_cases:
            raise HTTPException(status_code=400, detail="No test cases found in test_cases.xlsx. Please create test cases first.")
        
        # Analyze the HAR file
        calls = parse_har_for_calls(har_data, test_cases)
        url_failures, dimension_failures = analyze_failures(calls)
        test_groups = load_test_groups_from_json('../test_groups.json')
        group_results = evaluate_test_groups(calls, test_groups)

        # Create detailed results
        detailed_results = []
        for call in calls:
            for result, param_name, test_name in call.get("results", []):
                detailed_results.append(TestResult(
                    url=call["url"],
                    parameter=param_name,
                    result="Pass" if "Pass" in result else "Fail",
                    details=result,
                    test_case_name=test_name
                ))

        # Add group test results
        for name, data in group_results.items():
            detailed_results.append(TestResult(
                url="[GROUP]",
                parameter="",
                result="Pass" if data.get("passed") else "Fail",
                details=data.get("message", ""),
                test_case_name=name
            ))

        # Add group test results
        for name, data in group_results.items():
            detailed_results.append(TestResult(
                url="[GROUP]",
                parameter="",
                result="Pass" if data.get("passed") else "Fail",
                details=data.get("message", ""),
                test_case_name=name
            ))
        
        # Create analysis report
        total_tests = len(detailed_results)
        passed_tests = len([r for r in detailed_results if r.result == "Pass"])
        failed_tests = total_tests - passed_tests
        
        report = AnalysisReport(
            total_requests=len(calls),
            total_tests=total_tests,
            passed_tests=passed_tests,
            failed_tests=failed_tests,
            url_failures=url_failures,
            dimension_failures=dimension_failures,
            detailed_results=detailed_results,
            raw_data=calls,
            group_results=group_results
        )
        
        # Store results
        report_id = str(uuid.uuid4())
        analysis_results_store[report_id] = report
        
        return {"report_id": report_id, "report": report}
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid HAR file format")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")

@app.post("/api/analyze-har")
async def analyze_har(file: UploadFile = File(...)):
    if not file.filename.endswith('.har'):
        raise HTTPException(status_code=400, detail="Please upload a HAR file")
    
    try:
        # Read and parse HAR file
        contents = await file.read()
        har_data = json.loads(contents.decode('utf-8'))
        
        # Get test cases from Excel
        test_cases = load_tests_from_xlsx('../test_cases.xlsx')
        if not test_cases:
            raise HTTPException(status_code=400, detail="No test cases found in test_cases.xlsx. Please create test cases first.")
        
        # Analyze the HAR file
        calls = parse_har_for_calls(har_data, test_cases)
        url_failures, dimension_failures = analyze_failures(calls)
        test_groups = load_test_groups_from_json("../test_groups.json")
        group_results = evaluate_test_groups(calls, test_groups)
        
        # Create detailed results
        detailed_results = []
        for call in calls:
            for result, param_name, test_name in call.get("results", []):
                detailed_results.append(TestResult(
                    url=call["url"],
                    parameter=param_name,
                    result="Pass" if "Pass" in result else "Fail",
                    details=result,
                    test_case_name=test_name
                ))
        
        # Create analysis report
        total_tests = len(detailed_results)
        passed_tests = len([r for r in detailed_results if r.result == "Pass"])
        failed_tests = total_tests - passed_tests
        
        report = AnalysisReport(
            total_requests=len(calls),
            total_tests=total_tests,
            passed_tests=passed_tests,
            failed_tests=failed_tests,
            url_failures=url_failures,
            dimension_failures=dimension_failures,
            detailed_results=detailed_results,
            raw_data=calls,
            group_results=group_results
        )
        
        # Store results
        report_id = str(uuid.uuid4())
        analysis_results_store[report_id] = report
        
        return {"report_id": report_id, "report": report}
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid HAR file format")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")

@app.get("/api/reports/{report_id}")
async def get_report(report_id: str):
    if report_id not in analysis_results_store:
        raise HTTPException(status_code=404, detail="Report not found")
    return analysis_results_store[report_id]

@app.get("/api/reports")
async def get_all_reports():
    return {"reports": list(analysis_results_store.keys())}

@app.post("/api/reports/{report_id}/export")
async def export_report(report_id: str):
    if report_id not in analysis_results_store:
        raise HTTPException(status_code=404, detail="Report not found")
    
    report = analysis_results_store[report_id]
    
    # Create Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Test Results"
    
    # Headers
    headers = ["URL", "Parameter", "Result", "Details", "Test Case"]
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    
    # Data
    for row, result in enumerate(report.detailed_results, 2):
        sheet.cell(row=row, column=1, value=result.url)
        sheet.cell(row=row, column=2, value=result.parameter)
        sheet.cell(row=row, column=3, value=result.result)
        sheet.cell(row=row, column=4, value=result.details)
        sheet.cell(row=row, column=5, value=result.test_case_name)
    
    # Summary sheet
    summary_sheet = workbook.create_sheet("Summary")
    summary_data = [
        ["Metric", "Value"],
        ["Total Requests", report.total_requests],
        ["Total Tests", report.total_tests],
        ["Passed Tests", report.passed_tests],
        ["Failed Tests", report.failed_tests],
        ["", ""],
        ["URL Failures", "Count"]
    ]
    
    for url, count in report.url_failures.items():
        summary_data.append([url, count])
    
    summary_data.extend([["", ""], ["Parameter Failures", "Count"]])
    for param, count in report.dimension_failures.items():
        summary_data.append([param, count])

    summary_data.extend([["", ""], ["Test Groups", "Result"]])
    for name, data in report.group_results.items():
        status = "Pass" if data.get("passed") else "Fail"
        summary_data.append([name, f"{status}: {data.get('message', '')}"])
    
    for row, data in enumerate(summary_data, 1):
        for col, value in enumerate(data, 1):
            summary_sheet.cell(row=row, column=col, value=value)
    
    # Save to temp file
    temp_file = f"/tmp/harmony_report_{report_id}.xlsx"
    workbook.save(temp_file)
    
    return FileResponse(
        temp_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"harmony_report_{report_id}.xlsx"
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)