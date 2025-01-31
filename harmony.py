import json
import openpyxl
from collections import OrderedDict
from helper_functions import (
    find_har_file,
    load_har_file,
    load_tests_from_xlsx,
    parse_har_for_calls,
    combine_nested_keys,
)
from colorama import init

# Initialize colorama
init(autoreset=True)

def analyze_failures(calls):
    """Analyze and count failures based on call results."""
    url_failures = {}
    dimension_failures = {}
    
    for call in calls:
        for result, param_name in call.get("results", []):
            if "Fail" in result:
                url_failures[call['url']] = url_failures.get(call['url'], 0) + 1
                dimension_failures[param_name] = dimension_failures.get(param_name, 0) + 1
    
    return url_failures, dimension_failures

def write_results_to_excel(url_failures, dimension_failures, output_file='harmony_test_results.xlsx'):
    """Write URL and dimension failures to an Excel file."""
    workbook = openpyxl.load_workbook('template.xlsx')
    sheet = workbook.active
    
    # Clear existing data in the Test Results sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

    # Write URL Failures
    sheet['D1'], sheet['E1'] = 'URL Failures', 'Count'
    for row_num, (url, count) in enumerate(url_failures.items(), start=2):
        sheet.cell(row=row_num, column=4, value=url)
        sheet.cell(row=row_num, column=5, value=count)
    
    # Write Dimension Failures
    sheet['G1'], sheet['H1'] = 'Dimension Failures', 'Count'
    for row_num, (dimension, count) in enumerate(dimension_failures.items(), start=2):
        sheet.cell(row=row_num, column=7, value=dimension)
        sheet.cell(row=row_num, column=8, value=count)
    
    workbook.save(output_file)
    # print(f"Results saved to {output_file}")

def main():
    """Main execution function to process HAR file and test cases."""
    har_file_path = find_har_file()
    if not har_file_path:
        print("No HAR file found. Exiting.")
        return
    
    har_data = load_har_file(har_file_path)
    test_cases = load_tests_from_xlsx('test_cases.xlsx')
    
    adobe_calls = parse_har_for_calls(har_data, test_cases)
    url_failures, dimension_failures = analyze_failures(adobe_calls)
    
    write_results_to_excel(url_failures, dimension_failures)
    
    # Print results to console
    # for call in adobe_calls:
    # print(f"\nURL: {call['url']}")
    # print("Parameters:")
    # print(json.dumps(combine_nested_keys(call['parameters']), indent=2))
    # print("\nPayload:")
    # print("No payload" if call['payload'] == "No payload" else json.dumps(call['payload'], indent=2))
    # print("\nResults:")
    # for result in call.get("results", []):
    #     print(result)
    # print("-" * 50)

if __name__ == "__main__":
    main()